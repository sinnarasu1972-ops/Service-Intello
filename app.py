# app.py
import os
import io
import threading
import webbrowser
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict

from flask import Flask, jsonify, request, send_file, render_template_string
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

# =============================================================================
# 1) CONFIGURATION (LOCAL + RENDER SAFE)
# =============================================================================
# Render provides PORT env var. Locally you can still run with default 5000.
HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", "5000"))

# Auto-open browser only for local Windows usage (disabled on Render)
AUTO_OPEN_BROWSER = os.environ.get("AUTO_OPEN_BROWSER", "0").strip() in {"1", "true", "True", "yes", "YES"}

# Excel files:
# - On Render: keep files in repo under ./data OR set env vars PERSONAL_XLSX/MEAL_XLSX/...
# - On Windows local: you can still set env vars to D:\\... paths if you want.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "data"))

PERSONAL_XLSX = os.environ.get("PERSONAL_XLSX", os.path.join(DATA_DIR, "Link Triggered.xlsx"))
MEAL_XLSX = os.environ.get("MEAL_XLSX", os.path.join(DATA_DIR, "Link Triggered MEAL.xlsx"))
BODYSHOP_XLSX = os.environ.get("BODYSHOP_XLSX", os.path.join(DATA_DIR, "Link Triggered BP.xlsx"))
COMMERCIAL_XLSX = os.environ.get("COMMERCIAL_XLSX", os.path.join(DATA_DIR, "Link Triggered Commercial.xlsx"))

# FY Month order (Apr -> Mar)
MONTH_ORDER = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# =============================================================================
# 2) HELPERS / UTILITIES
# =============================================================================
def s(v: Any) -> Optional[str]:
    if v is None:
        return None
    t = str(v).strip()
    if not t:
        return None
    if t == "-":
        return None
    return t


def to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)

    t = str(v).strip()
    if not t:
        return None
    if t.lower() in {"na", "n/a", "none", "-"}:
        return None

    t = t.replace(",", "").replace("₹", "").strip()

    if t.endswith("%"):
        t2 = t[:-1].strip()
        try:
            return float(t2)
        except Exception:
            return None

    try:
        return float(t)
    except Exception:
        return None


def r2(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except Exception:
        return None


def safe_div(num: Optional[float], den: Optional[float]) -> Optional[float]:
    if num is None or den is None or den == 0:
        return None
    return num / den


def normalize_header(h: Any) -> str:
    if h is None:
        return ""
    t = str(h).strip()
    if not t:
        return ""
    t = " ".join(t.split())
    if t.lower() == "diviosion":
        return "Division"
    return t


def detect_month(sheet_name: str) -> Optional[str]:
    if not sheet_name:
        return None
    key = sheet_name.strip()[:3].lower()
    for m in MONTH_ORDER:
        if m.lower() == key:
            return m
    return None


def avg(values: List[Optional[float]]) -> Optional[float]:
    nums = [v for v in values if isinstance(v, (int, float)) and v is not None]
    if not nums:
        return None
    return sum(nums) / len(nums)


def export_xlsx(columns: List[str], rows: List[Dict[str, Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(columns)
    for row in rows:
        ws.append(["" if row.get(c) is None else row.get(c) for c in columns])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def key_norm(x: Optional[str]) -> str:
    return (x or "All").strip().lower()


# =============================================================================
# 3) COLUMN DETECTION (STRONGER)
# =============================================================================
def build_header_map(header_row: List[Any]) -> Dict[str, int]:
    hmap: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        nh = normalize_header(h)
        if nh:
            hmap[nh.lower()] = i
    return hmap


def find_col_index_any(hmap: Dict[str, int], tokens: List[str]) -> Optional[int]:
    for k, idx in hmap.items():
        kk = (k or "").lower()
        for t in tokens:
            if t in kk:
                return idx
    return None


def find_col_index_priority(hmap: Dict[str, int], token_groups: List[List[str]]) -> Optional[int]:
    for tokens in token_groups:
        idx = find_col_index_any(hmap, tokens)
        if idx is not None:
            return idx
    return None


# =============================================================================
# 4) DATASET
# =============================================================================
class Dataset:
    def __init__(self, name: str, excel_path: str):
        self.name = name
        self.excel_path = excel_path
        self.data_rows: List[Dict[str, Any]] = []
        self.available_months: List[str] = []
        self.load_error: Optional[str] = None

        self.index: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        self.div_by_month: Dict[str, set] = defaultdict(set)
        self.sa_by_month_div: Dict[Tuple[str, str], set] = defaultdict(set)

        self._load_and_index()

    def _load_and_index(self) -> None:
        try:
            self.data_rows, self.available_months = self._load_excel()
            self._build_indexes(self.data_rows)
            self.load_error = None
        except Exception as e:
            self.load_error = str(e)
            self.data_rows = []
            self.available_months = []
            self._build_indexes([])

    def _load_excel(self) -> Tuple[List[Dict[str, Any]], List[str]]:
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        wb = load_workbook(self.excel_path, data_only=True)
        all_rows: List[Dict[str, Any]] = []
        months_present = set()

        for sheet in wb.sheetnames:
            month = detect_month(sheet)
            if not month:
                continue
            months_present.add(month)

            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)

            try:
                header = list(next(it))
            except StopIteration:
                continue

            hmap = build_header_map(header)

            c_mile = find_col_index_priority(hmap, [["mile id"], ["mileid"], ["mile"], ["mile_id"]])
            c_links = find_col_index_priority(hmap, [["links triggered"], ["links trig"], ["link triggered"], ["links"], ["trigger"]])
            c_resp = find_col_index_priority(hmap, [["total response"], ["responses"], ["response"], ["respon"], ["reply"]])
            c_nps = find_col_index_priority(hmap, [["nps"], ["np score"], ["np"]])
            c_concern = find_col_index_priority(hmap, [["concern count"], ["concern"], ["complaint"]])
            c_cc = find_col_index_priority(hmap, [["cc/1000"], ["cc/10"], ["cc per"], ["cc per 1000"]])
            c_osat = find_col_index_priority(hmap, [["osat"], ["overall satisfaction"], ["overall sat"], ["osat%"], ["os"]])
            c_sa = find_col_index_priority(hmap, [["sa name"], ["service advisor name"], ["service advisor"], ["advisor name"], ["advisor"], ["sa  "], ["sa_"], ["sa"]])
            c_div = find_col_index_priority(
                hmap,
                [
                    ["division name"],
                    ["division"],
                    ["diviosion"],
                    ["divn"],
                    ["div."],
                    ["div "],
                    ["div"],
                    ["branch name"],
                    ["branch"],
                    ["outlet"],
                    ["workshop"],
                    ["dealer location"],
                    ["location"],
                ],
            )

            def cell(row: Tuple[Any, ...], idx: Optional[int]) -> Any:
                if idx is None:
                    return None
                if idx >= len(row):
                    return None
                return row[idx]

            for row in it:
                if not row:
                    continue

                links = to_float(cell(row, c_links))
                resp = to_float(cell(row, c_resp))
                nps = to_float(cell(row, c_nps))
                concern = to_float(cell(row, c_concern))
                cc = to_float(cell(row, c_cc))

                osat = to_float(cell(row, c_osat))
                if osat is not None and 0 <= osat <= 1:
                    osat = r2(osat * 100.0)
                else:
                    osat = r2(osat)

                pct = None
                ratio = safe_div(resp, links)
                if ratio is not None:
                    pct = r2(ratio * 100.0)

                rec = {
                    "Month": month,
                    "SA Name": s(cell(row, c_sa)),
                    "Division": s(cell(row, c_div)),
                    "Mile id": s(cell(row, c_mile)),
                    "Links Triggered": links,
                    "Response": resp,
                    "NPS": nps,
                    "% of Response": pct,
                    "Concern Count": concern,
                    "CC/1000": cc,
                    "OSAT": osat,
                }

                if (
                    not rec["SA Name"]
                    and not rec["Division"]
                    and rec["Links Triggered"] is None
                    and rec["Response"] is None
                    and rec["NPS"] is None
                    and rec["OSAT"] is None
                ):
                    continue

                all_rows.append(rec)

        available_months = [m for m in MONTH_ORDER if m in months_present]
        return all_rows, available_months

    def _build_indexes(self, rows: List[Dict[str, Any]]) -> None:
        self.index.clear()
        self.div_by_month.clear()
        self.sa_by_month_div.clear()

        rows.sort(key=lambda r: ((r.get("Division") or ""), (r.get("SA Name") or ""), (r.get("Month") or "")))

        for r in rows:
            m = key_norm(r.get("Month"))
            d = key_norm(r.get("Division"))
            a = key_norm(r.get("SA Name"))

            m_txt = (r.get("Month") or "").strip()
            d_txt = (r.get("Division") or "").strip()
            a_txt = (r.get("SA Name") or "").strip()

            if m_txt and d_txt:
                self.div_by_month[m].add(d_txt)
            if m_txt and d_txt and a_txt:
                self.sa_by_month_div[(m, d)].add(a_txt)

            # base
            self.index[(m, d, a)].append(r)

            # add All-keys
            self.index[(m, d, "all")].append(r)
            self.index[(m, "all", a)].append(r)
            self.index[(m, "all", "all")].append(r)

            self.index[("all", d, a)].append(r)
            self.index[("all", d, "all")].append(r)

            self.index[("all", "all", a)].append(r)
            self.index[("all", "all", "all")].append(r)

    def apply_filters(self, month: str, division: str, sa_name: str) -> List[Dict[str, Any]]:
        # Supports:
        # - "All"
        # - "__NONE__" (means user deselected all -> return empty)
        # - comma-separated multi selections like "Apr,May"
        def parse_sel(v: str) -> List[str]:
            t = (v or "All").strip()
            if not t:
                return ["all"]
            if t == "__NONE__":
                return []
            if t.lower() == "all":
                return ["all"]
            parts = [p.strip() for p in t.split(",") if p.strip()]
            return parts if parts else ["all"]

        months = parse_sel(month)
        divs = parse_sel(division)
        sas = parse_sel(sa_name)

        if not months or not divs or not sas:
            return []

        out: List[Dict[str, Any]] = []
        seen = set()

        for mo in months:
            for di in divs:
                for sa_ in sas:
                    key = (key_norm(mo), key_norm(di), key_norm(sa_))
                    rows = self.index.get(key, [])
                    for r in rows:
                        rid = id(r)
                        if rid in seen:
                            continue
                        seen.add(rid)
                        out.append(r)
        return out

    def compute_kpis(self, rows: List[Dict[str, Any]], include_osat: bool = True) -> Dict[str, Any]:
        total_links = sum((r.get("Links Triggered") or 0) for r in rows if r.get("Links Triggered") is not None)
        total_resp = sum((r.get("Response") or 0) for r in rows if r.get("Response") is not None)
        total_concern = sum((r.get("Concern Count") or 0) for r in rows if r.get("Concern Count") is not None)

        avg_pct = r2((total_resp / total_links) * 100.0) if total_links else None
        cc_1000 = r2((total_concern / total_links) * 1000.0) if total_links else None

        out = {
            "total_links_triggered": int(total_links),
            "total_responses": int(total_resp),
            "avg_percent_response": avg_pct,
            "total_concern_count": int(total_concern),
            "avg_cc_per_1000": cc_1000,
            "record_count": len(rows),
        }

        if include_osat:
            out["avg_osat"] = r2(avg([r.get("OSAT") for r in rows]))

        return out

    def get_filters(self, sel_month: str, sel_div: str) -> Dict[str, List[str]]:
        if self.load_error:
            raise RuntimeError(self.load_error)

        def parse_sel(v: str) -> List[str]:
            t = (v or "All").strip()
            if not t:
                return ["all"]
            if t == "__NONE__":
                return []
            if t.lower() == "all":
                return ["all"]
            parts = [p.strip() for p in t.split(",") if p.strip()]
            return parts if parts else ["all"]

        months = parse_sel(sel_month)
        divs = parse_sel(sel_div)

        if not months:
            return {"months": self.available_months or MONTH_ORDER, "divisions": [], "sa_names": []}

        # Divisions based on selected months
        if "all" in [m.lower() for m in months]:
            all_divs = set()
            for ds in self.div_by_month.values():
                all_divs |= set(ds)
            divisions = sorted(all_divs)
        else:
            div_set = set()
            for mo in months:
                div_set |= set(self.div_by_month.get(key_norm(mo), set()))
            divisions = sorted(div_set)

        # SA Names
        if not divs:
            return {"months": self.available_months or MONTH_ORDER, "divisions": divisions, "sa_names": []}

        want_all_div = "all" in [d.lower() for d in divs]

        if "all" in [m.lower() for m in months] and want_all_div:
            sa_names = sorted({(r.get("SA Name") or "").strip() for r in self.data_rows if (r.get("SA Name") or "").strip()})
            return {"months": self.available_months or MONTH_ORDER, "divisions": divisions, "sa_names": sa_names}

        sa_set = set()

        if "all" in [m.lower() for m in months]:
            want_divs = {key_norm(d) for d in divs if d.lower() != "all"}
            for r in self.data_rows:
                if (r.get("SA Name") or "").strip() and key_norm(r.get("Division")) in want_divs:
                    sa_set.add((r.get("SA Name") or "").strip())
        else:
            for mo in months:
                mm = key_norm(mo)
                if want_all_div:
                    for (m_key, _d_key), names in self.sa_by_month_div.items():
                        if m_key == mm:
                            sa_set |= set(names)
                else:
                    for dv in divs:
                        dd = key_norm(dv)
                        sa_set |= set(self.sa_by_month_div.get((mm, dd), set()))

        sa_names = sorted(sa_set)
        return {"months": self.available_months or MONTH_ORDER, "divisions": divisions, "sa_names": sa_names}


PERSONAL = Dataset("Personal", PERSONAL_XLSX)
MEAL = Dataset("MEAL", MEAL_XLSX)
BODYSHOP = Dataset("Body Shop", BODYSHOP_XLSX)
COMMERCIAL = Dataset("Commercial", COMMERCIAL_XLSX)

DATASETS: Dict[str, Dataset] = {
    "personal": PERSONAL,
    "meal": MEAL,
    "bodyshop": BODYSHOP,
    "commercial": COMMERCIAL,
}


def get_dataset(key: str) -> Dataset:
    return DATASETS.get((key or "").strip().lower(), PERSONAL)


DATASET_META = {
    "personal": {
        "page_title": "UNNATI MOTORS - SERVICE INTELLO (PERSONAL)",
        "show_osat": True,
        "table_mode": "normal",
        "export_cols": [
            "Month",
            "SA Name",
            "Links Triggered",
            "Response",
            "% of Response",
            "Concern Count",
            "CC/1000",
            "OSAT",
            "NPS",
            "Division",
            "Mile id",
        ],
    },
    "meal": {
        "page_title": "UNNATI MOTORS - SERVICE INTELLO (MEAL)",
        "show_osat": True,
        "table_mode": "normal",
        "export_cols": [
            "Month",
            "SA Name",
            "Links Triggered",
            "Response",
            "% of Response",
            "Concern Count",
            "CC/1000",
            "OSAT",
            "NPS",
            "Division",
            "Mile id",
        ],
    },
    "bodyshop": {
        "page_title": "UNNATI MOTORS - SERVICE INTELLO (BODY SHOP)",
        "show_osat": False,
        "table_mode": "bodyshop",
        "export_cols": [
            "Month",
            "SA Name",
            "Links Triggered",
            "Response",
            "% of Response",
            "Concern Count",
            "CC/1000",
            "Division",
            "Mile id",
        ],
    },
    "commercial": {
        "page_title": "UNNATI MOTORS - SERVICE INTELLO (COMMERCIAL)",
        "show_osat": True,
        "table_mode": "normal",
        "export_cols": [
            "Month",
            "SA Name",
            "Links Triggered",
            "Response",
            "% of Response",
            "Concern Count",
            "CC/1000",
            "OSAT",
            "NPS",
            "Division",
            "Mile id",
        ],
    },
}


# =============================================================================
# 6) HTML (ALL IN ONE)
# =============================================================================
INDEX_HTML = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{{ page_title }}</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <style>
    :root{
      --navy:#5b62d6;
      --navy2:#7a5cf0;
      --bg:#f3f5ff;
      --card:#ffffff;
      --text:#0f172a;
      --muted:#64748b;
      --border:#e6e8ff;
      --shadow: 0 10px 25px rgba(0,0,0,0.10);
      --radius:14px;
      --accent:#f97316;
    }
    *{box-sizing:border-box;}
    body{margin:0;font-family:Poppins,system-ui,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);color:var(--text);}

    .header{
      background: linear-gradient(90deg, var(--navy) 0%, var(--navy2) 100%);
      color:#fff;padding:18px 16px;box-shadow: var(--shadow);
    }
    .header .wrap{max-width:1200px;margin:0 auto;}
    .title{margin:0;font-size:26px;font-weight:800;text-align:center;}
    .subtitle{margin-top:6px;text-align:center;color:rgba(255,255,255,0.85);font-size:13px;font-weight:500;}

    .tabs-wrap{max-width:1200px;margin:12px auto 0; padding:0 12px;}
    .tabs{display:flex;gap:10px;flex-wrap:wrap;justify-content:center;}
    .tab{
      text-decoration:none;
      display:inline-flex;
      align-items:center;
      justify-content:center;
      padding:10px 18px;
      border-radius:10px;
      border:1px solid rgba(255,255,255,0.35);
      color:#fff;
      font-weight:800;
      background: rgba(255,255,255,0.15);
      min-width:120px;
    }
    .tab.active{
      background: rgba(255,255,255,0.95);
      color: #1f2a6b;
      border-color: rgba(255,255,255,0.95);
    }

    .container{max-width:1200px;margin:16px auto 26px; padding:0 12px;}
    .filters{background:var(--card);border:1px solid var(--border);border-radius: var(--radius);padding:14px;box-shadow: var(--shadow);}
    .filter-row{display:grid;grid-template-columns: 1fr 1fr 1fr auto;gap:12px;align-items:end;}
    @media (max-width: 900px){.filter-row{grid-template-columns: 1fr 1fr;}.reset-btn{grid-column: 1 / -1;}}
    label{display:block;font-size:12px;color:var(--muted);font-weight:800;margin-bottom:6px;}
    select{width:100%;padding:10px 12px;border-radius:12px;border:1px solid var(--border);background:#fff;outline:none;font-family:inherit;font-size:13px;}
    .reset-btn{border:none;padding:10px 14px;border-radius:10px;cursor:pointer;font-weight:900;background:#ef4444;color:#fff;height:42px;white-space:nowrap;}

    /* ===== Multi-select dropdown (Month/Division) ===== */
    .ms{position:relative;width:100%;}
    .ms-btn{
      width:100%;
      padding:10px 12px;
      border-radius:12px;
      border:1px solid var(--border);
      background:#fff;
      outline:none;
      font-family:inherit;
      font-size:13px;
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:10px;
      cursor:pointer;
      min-height:42px;
    }
    .ms-btn .left{display:flex; align-items:center; gap:10px; min-width:0;}
    .ms-btn .label{
      font-weight:700;color: var(--text);
      white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
    }
    .ms-btn .badge{
      background: rgba(249,115,22,0.12);
      color: var(--accent);
      border: 1px solid rgba(249,115,22,0.25);
      font-weight:900;
      padding:3px 8px;
      border-radius:999px;
      font-size:11px;
      flex:0 0 auto;
    }
    .ms-btn .caret{color: var(--muted); font-weight:900; flex:0 0 auto;}
    .ms-panel{
      position:absolute;top:calc(100% + 8px);left:0;right:0;
      background:#fff;border:1px solid var(--border);
      border-radius:12px;box-shadow: var(--shadow);
      z-index:50;display:none;overflow:hidden;
    }
    .ms-panel.open{display:block;}
    .ms-panel .top{
      padding:10px 12px;border-bottom:1px solid var(--border);
      display:flex;align-items:center;justify-content:space-between;gap:10px;
    }
    .ms-panel .top .title{font-size:12px;font-weight:900;color: var(--muted);margin:0;}
    .ms-panel .list{
      max-height:260px;overflow:auto;padding:10px 12px;
      display:flex;flex-direction:column;gap:8px;
    }
    .ms-item{display:flex;align-items:center;gap:10px;font-size:12.5px;color: var(--text);font-weight:700;user-select:none;}
    .ms-item input{transform: translateY(1px);}

    .kpis{margin-top:14px;display:grid;grid-template-columns: repeat(6, minmax(0,1fr));gap:12px;}
    @media (max-width: 1050px){ .kpis{grid-template-columns: repeat(3, minmax(0,1fr));} }
    @media (max-width: 650px){ .kpis{grid-template-columns: repeat(2, minmax(0,1fr));} }
    .kpi{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:12px;box-shadow: var(--shadow);min-height:92px;display:flex;flex-direction:column;justify-content:center;gap:6px;}
    .kpi .klabel{font-size:11px;color:var(--muted);font-weight:900;text-transform:uppercase;letter-spacing:.4px;}
    .kpi .kvalue{font-size:22px;font-weight:900;}
    .kpi.blue{border-left:6px solid #3b82f6;}
    .kpi.green{border-left:6px solid #22c55e;}
    .kpi.teal{border-left:6px solid #14b8a6;}
    .kpi.orange{border-left:6px solid #f97316;}
    .kpi.red{border-left:6px solid #ef4444;}
    .kpi.purple{border-left:6px solid #a855f7;}

    /* ===== Chart Card ===== */
    .chartcard{
      margin-top:14px;background:var(--card);
      border:1px solid var(--border);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding:12px 14px 14px;
      display:none;
    }
    .chartcard.show{display:block;}
    .charttitle{font-weight:900;color: var(--text);font-size:13px;margin:0 0 10px 0;}

    .chartwrap{position:relative;}
    canvas{width:100% !important; height:280px !important;}
    .ptLabel{
      position:absolute;
      transform: translate(-50%, -135%);
      font-weight:900;
      font-size:11px;
      color:#0f172a;
      background: rgba(255,255,255,0.90);
      border:1px solid rgba(15,23,42,0.10);
      padding:2px 6px;
      border-radius:999px;
      pointer-events:none;
      white-space:nowrap;
    }

    /* ===== Table ===== */
    .tablecard{margin-top:14px;background:var(--card);border:1px solid var(--border);border-radius: var(--radius);box-shadow: var(--shadow);overflow:hidden;}
    .tablehead{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:12px 14px;border-bottom:1px solid var(--border);}
    .tablehead .left{display:flex;flex-direction:column;gap:2px;}
    .tablehead h3{margin:0;font-size:14px;font-weight:900;}
    .count{font-size:12px;color:var(--muted);font-weight:700;}
    .table-actions{display:flex;align-items:center;gap:10px;flex-wrap:wrap;}
    .btn{border:1px solid var(--border);background:#fff;border-radius:10px;padding:9px 12px;cursor:pointer;font-weight:900;font-family:inherit;font-size:12px;}
    .btn.primary{background:#0ea5e9;color:#fff;border-color:#0ea5e9;}
    .btn.gray{background:#eef2ff;}
    .pager{display:flex;align-items:center;gap:8px;}
    .pager .btn{padding:7px 10px;}
    .pill-green{color:#16a34a;font-weight:900;}
    .pill-red{color:#ef4444;font-weight:900;}

    table{width:100%;border-collapse:collapse;}
    th, td{padding:10px 10px;border-bottom:1px solid var(--border);text-align:left;font-size:12.5px;vertical-align:top;}
    th{background:#f8fafc;font-weight:900;color:#0f172a;position:sticky;top:0;z-index:1;}
    tr:hover td{background:#fbfdff;}

    .loading{display:none;margin-top:12px;padding:10px 12px;background:#fff;border:1px dashed var(--border);border-radius:12px;color:var(--muted);font-weight:800;}
    .loading.show{display:block;}
    .err{display:none;margin-top:12px;padding:10px 12px;background:#fff;border:1px solid #fecaca;border-radius:12px;color:#b91c1c;font-weight:900;}
    .err.show{display:block;}

    /* Column filter pop */
    .pop{position:fixed;inset:0;display:none;align-items:center;justify-content:center;background:rgba(2,6,23,0.35);z-index:100;}
    .pop.show{display:flex;}
    .pop .box{width:min(560px,92vw);background:#fff;border-radius:16px;box-shadow: var(--shadow);border:1px solid var(--border);overflow:hidden;}
    .pop .hd{padding:12px 14px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;}
    .pop .hd b{font-size:13px;}
    .pop .bd{padding:12px 14px;}
    .pop input[type="text"]{width:100%;padding:10px 12px;border-radius:12px;border:1px solid var(--border);font-family:inherit;}
    .pop .list{margin-top:10px;max-height:260px;overflow:auto;border:1px solid var(--border);border-radius:12px;padding:10px;}
    .pop .ft{padding:12px 14px;border-top:1px solid var(--border);display:flex;justify-content:flex-end;gap:10px;}
    .pop .item{display:flex;gap:10px;align-items:center;font-size:12.5px;font-weight:700;margin:6px 0;}
  </style>
</head>
<body>
  <div class="header">
    <div class="wrap">
      <h1 class="title">{{ page_title }}</h1>
      <div class="subtitle">Month & Division multi-select • Deselect All works • CC/1000 month trend</div>
    </div>

    <div class="tabs-wrap">
      <div class="tabs">
        <a class="tab {{ 'active' if active_key=='personal' else '' }}" href="/personal">PERSONAL</a>
        <a class="tab {{ 'active' if active_key=='meal' else '' }}" href="/meal">MEAL</a>
        <a class="tab {{ 'active' if active_key=='bodyshop' else '' }}" href="/bodyshop">BODY SHOP</a>
        <a class="tab {{ 'active' if active_key=='commercial' else '' }}" href="/commercial">COMMERCIAL</a>
      </div>
    </div>
  </div>

  <div class="container">

    <div class="filters">
      <div class="filter-row">
        <!-- Month Multi Select -->
        <div>
          <label>Month (Multi Select)</label>
          <div class="ms" id="msMonth">
            <button class="ms-btn" id="msMonthBtn">
              <div class="left">
                <span class="label" id="msMonthLabel">All</span>
                <span class="badge" id="msMonthBadge" style="display:none;">0</span>
              </div>
              <span class="caret">▾</span>
            </button>
            <div class="ms-panel" id="msMonthPanel">
              <div class="top">
                <p class="title">Select Months</p>
                <label class="ms-item"><input type="checkbox" id="msMonthAll"> <span>All / Deselect</span></label>
              </div>
              <div class="list" id="msMonthList"></div>
            </div>
          </div>
          <input type="hidden" id="month" value="All">
        </div>

        <!-- Division Multi Select -->
        <div>
          <label>Division (Multi Select)</label>
          <div class="ms" id="msDiv">
            <button class="ms-btn" id="msDivBtn">
              <div class="left">
                <span class="label" id="msDivLabel">All</span>
                <span class="badge" id="msDivBadge" style="display:none;">0</span>
              </div>
              <span class="caret">▾</span>
            </button>
            <div class="ms-panel" id="msDivPanel">
              <div class="top">
                <p class="title">Select Divisions</p>
                <label class="ms-item"><input type="checkbox" id="msDivAll"> <span>All / Deselect</span></label>
              </div>
              <div class="list" id="msDivList"></div>
            </div>
          </div>
          <input type="hidden" id="division" value="All">
        </div>

        <div>
          <label>SA Name</label>
          <select id="sa">
            <option value="All">All</option>
          </select>
        </div>

        <button class="reset-btn" id="reset">Reset</button>
      </div>
    </div>

    <div class="loading" id="loading">Loading…</div>
    <div class="err" id="err"></div>

    <div class="kpis">
      <div class="kpi blue"><div class="klabel">Links Triggered</div><div class="kvalue" id="k1">-</div></div>
      <div class="kpi green"><div class="klabel">Responses</div><div class="kvalue" id="k2">-</div></div>
      <div class="kpi teal"><div class="klabel">Avg % Response</div><div class="kvalue" id="k3">-</div></div>
      <div class="kpi orange"><div class="klabel">Concern Count</div><div class="kvalue" id="k4">-</div></div>
      <div class="kpi red"><div class="klabel">CC/1000</div><div class="kvalue" id="k5">-</div></div>
      <div class="kpi purple"><div class="klabel">OSAT</div><div class="kvalue" id="k6">-</div></div>
    </div>

    <div class="chartcard" id="chartCard">
      <p class="charttitle">Month-Wise CC/1000 Trend</p>
      <div class="chartwrap" id="chartWrap">
        <canvas id="ccChart"></canvas>
      </div>
    </div>

    <div class="tablecard">
      <div class="tablehead">
        <div class="left">
          <h3>Records</h3>
          <div class="count" id="count"><span>Showing 0 records</span></div>
        </div>
        <div class="table-actions">
          <button class="btn gray" id="clearColFilters">Clear Column Filters</button>
          <button class="btn primary" id="exportBtn">Export</button>
          <div class="pager">
            <button class="btn" id="prevBtn">Prev</button>
            <button class="btn" id="nextBtn">Next</button>
          </div>
          <select id="pageSize">
            <option value="25">25</option>
            <option value="50" selected>50</option>
            <option value="100">100</option>
            <option value="200">200</option>
          </select>
        </div>
      </div>

      <div style="overflow:auto; max-height:520px;">
        <table>
          <thead>
            {% if table_mode == "bodyshop" %}
            <tr>
              <th>Month</th>
              <th>SA Name</th>
              <th>Links Triggered</th>
              <th>Response</th>
              <th>% Response</th>
              <th>Concern Count</th>
              <th>CC/1000</th>
              <th>Division</th>
            </tr>
            {% else %}
            <tr>
              <th>Month</th>
              <th>SA Name</th>
              <th>Links Triggered</th>
              <th>Response</th>
              <th>% Response</th>
              <th>Concern Count</th>
              <th>CC/1000</th>
              <th>OSAT</th>
              <th>NPS</th>
              <th>Division</th>
            </tr>
            {% endif %}
          </thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Column filter popup -->
  <div class="pop" id="colFilterPop">
    <div class="box">
      <div class="hd">
        <b id="colFilterTitle">Filter</b>
        <button class="btn" id="colFilterCancel">X</button>
      </div>
      <div class="bd">
        <input type="text" id="colFilterSearch" placeholder="Search values...">
        <div class="list" id="colFilterList"></div>
        <label class="item"><input type="checkbox" id="colFilterAll"> <span>Select All</span></label>
      </div>
      <div class="ft">
        <button class="btn" id="colFilterOk">Apply</button>
      </div>
    </div>
  </div>

<script>
  const DATASET = "{{ active_key }}";
  const SHOW_OSAT = {{ "true" if show_osat else "false" }};
  const TABLE_MODE = "{{ table_mode }}";
  const FY_MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"];

  const DIV_MAP = {
    "NAGPUR_KAMPTHEE ROAD": "HO",
    "YAVATMAL": "YAT",
    "WAGHOLI": "WAG",
    "CHAUFULA_SZZ": "CHA",
    "AMRAVATI": "AMT",
    "KOLHAPUR_WS": "KOL",
    "SHIKRAPUR_SZS": "SHI",
    "CHIKHALI": "CHI",
    "NAGPUR_WARDHAMAN NGR": "CITY",
    "NAGPUR_WARDHAMANNGR": "CITY",
    "NAGPUR_WARDHAMAN_NGR": "CITY"
  };
  function divShort(name){
    const t = (name || "").trim();
    return DIV_MAP[t] || t;
  }

  const loading = document.getElementById("loading");
  const err = document.getElementById("err");
  const tbody = document.getElementById("tbody");
  const count = document.getElementById("count");

  const month = document.getElementById("month");
  const division = document.getElementById("division");
  const sa = document.getElementById("sa");
  const pageSize = document.getElementById("pageSize");

  const k1 = document.getElementById("k1");
  const k2 = document.getElementById("k2");
  const k3 = document.getElementById("k3");
  const k4 = document.getElementById("k4");
  const k5 = document.getElementById("k5");
  const k6 = document.getElementById("k6");

  const chartCard = document.getElementById("chartCard");
  const chartWrap = document.getElementById("chartWrap");
  const ccChartCanvas = document.getElementById("ccChart");
  let ccChart = null;
  let pointEls = [];

  // Multi-select objects
  const msMonth = {
    root: document.getElementById("msMonth"),
    btn: document.getElementById("msMonthBtn"),
    panel: document.getElementById("msMonthPanel"),
    list: document.getElementById("msMonthList"),
    all: document.getElementById("msMonthAll"),
    label: document.getElementById("msMonthLabel"),
    badge: document.getElementById("msMonthBadge"),
    selected: new Set(),
    options: [],
    inited: false
  };

  const msDiv = {
    root: document.getElementById("msDiv"),
    btn: document.getElementById("msDivBtn"),
    panel: document.getElementById("msDivPanel"),
    list: document.getElementById("msDivList"),
    all: document.getElementById("msDivAll"),
    label: document.getElementById("msDivLabel"),
    badge: document.getElementById("msDivBadge"),
    selected: new Set(),
    options: [],
    inited: false
  };

  let allRows = [];
  let lastFilteredRows = [];
  let page = 1;

  function showLoading(x){ x ? loading.classList.add("show") : loading.classList.remove("show"); }
  function safe(v){ if(v===null || v===undefined) return "-"; const t=String(v).trim(); return t ? t : "-"; }
  function num(v){ const n=Number(v); return Number.isFinite(n) ? n : null; }
  function pctText(v){ const n=num(v); return n===null ? "-" : n.toFixed(2) + "%"; }

  function params(){
    return new URLSearchParams({
      dataset: DATASET,
      month: month.value,
      division: division.value,
      sa_name: sa.value
    }).toString();
  }

  function monthDivParamsForFilters(){
    return new URLSearchParams({
      dataset: DATASET,
      month: month.value,
      division: division.value
    }).toString();
  }

  function openPanel(ms){ ms.panel.classList.add("open"); }
  function closePanel(ms){ ms.panel.classList.remove("open"); }
  function togglePanel(ms){ ms.panel.classList.contains("open") ? closePanel(ms) : openPanel(ms); }

  // IMPORTANT:
  // - If selected is empty -> send "__NONE__" so backend returns 0 records (Deselect All works correctly)
  function syncMultiSelectHiddenValue(ms, hiddenInput){
    if(ms.selected.size === 0){ hiddenInput.value = "__NONE__"; return; }
    if(ms.options.length === 0){ hiddenInput.value = "All"; return; }
    const allSelected = ms.selected.size === ms.options.length;
    hiddenInput.value = allSelected ? "All" : Array.from(ms.selected).join(",");
  }

  function syncMultiSelectUI(ms){
    const total = ms.options.length;
    const sel = ms.selected.size;

    if(total === 0){
      ms.label.textContent = "All";
      ms.badge.style.display = "none";
      ms.all.checked = true;
      ms.all.indeterminate = false;
      return;
    }

    const allSelected = sel === total;
    const noneSelected = sel === 0;

    if(allSelected){
      ms.label.textContent = "All";
      ms.badge.style.display = "none";
      ms.all.checked = true;
      ms.all.indeterminate = false;
    }else{
      ms.label.textContent = noneSelected ? "None" : "Selected";
      ms.badge.style.display = "inline-flex";
      ms.badge.textContent = sel + " Selected";
      ms.all.checked = false;
      ms.all.indeterminate = (!noneSelected && !allSelected);
    }

    ms.list.querySelectorAll("input[type='checkbox'][data-opt]").forEach(ch => {
      const opt = ch.getAttribute("data-opt");
      ch.checked = ms.selected.has(opt);
    });
  }

  function rebuildMultiSelect(ms, options, keepSelection=true){
    ms.options = Array.isArray(options) ? options.slice() : [];

    const prev = new Set(ms.selected);
    ms.selected.clear();

    if(keepSelection){
      ms.options.forEach(o => { if(prev.has(o)) ms.selected.add(o); });
    }

    // ONLY first init: default select all
    if(!ms.inited){
      if(ms.selected.size === 0 && ms.options.length > 0){
        ms.options.forEach(o => ms.selected.add(o));
      }
      ms.inited = true;
    }

    ms.list.innerHTML = "";
    ms.options.forEach(opt => {
      const display = (ms === msDiv) ? divShort(opt) : opt;
      const id = ms.root.id + "_" + opt.replace(/[^a-z0-9]/gi,'_');
      ms.list.insertAdjacentHTML("beforeend", `
        <label class="ms-item" for="${id}">
          <input type="checkbox" id="${id}" data-opt="${opt}">
          <span>${display}</span>
        </label>
      `);
    });

    ms.list.querySelectorAll("input[type='checkbox'][data-opt]").forEach(ch => {
      ch.addEventListener("change", async (e) => {
        const opt = e.target.getAttribute("data-opt");
        if(e.target.checked) ms.selected.add(opt);
        else ms.selected.delete(opt);

        syncMultiSelectUI(ms);
        if(ms === msMonth) syncMultiSelectHiddenValue(ms, month);
        if(ms === msDiv) syncMultiSelectHiddenValue(ms, division);

        await loadFilters();
        page = 1;
        await refresh();
      });
    });

    // select all / deselect all
    ms.all.onchange = async (e)=>{
      e.stopPropagation();
      const total = ms.options.length;
      const allSelected = ms.selected.size === total;

      if(allSelected){
        ms.selected.clear(); // deselect all -> will become "__NONE__"
      }else{
        ms.options.forEach(o => ms.selected.add(o)); // select all
      }

      syncMultiSelectUI(ms);
      if(ms === msMonth) syncMultiSelectHiddenValue(ms, month);
      if(ms === msDiv) syncMultiSelectHiddenValue(ms, division);

      await loadFilters();
      page = 1;
      await refresh();
    };

    syncMultiSelectUI(ms);
  }

  msMonth.btn.addEventListener("click", (e)=>{ e.preventDefault(); e.stopPropagation(); togglePanel(msMonth); closePanel(msDiv); });
  msDiv.btn.addEventListener("click", (e)=>{ e.preventDefault(); e.stopPropagation(); togglePanel(msDiv); closePanel(msMonth); });

  document.addEventListener("click", (e)=>{
    if(msMonth.panel.classList.contains("open") && !msMonth.root.contains(e.target)) closePanel(msMonth);
    if(msDiv.panel.classList.contains("open") && !msDiv.root.contains(e.target)) closePanel(msDiv);
  });

  async function loadFilters(){
    const res = await fetch("/api/filters?" + monthDivParamsForFilters());
    const j = await res.json();
    if(!j.ok) throw new Error(j.error || "Failed to load filters");

    rebuildMultiSelect(msMonth, j.filters.months || FY_MONTHS, true);
    rebuildMultiSelect(msDiv, j.filters.divisions || [], true);

    // SA dropdown
    const cur = sa.value || "All";
    sa.innerHTML = `<option value="All">All</option>` + (j.filters.sa_names || []).map(x=>`<option value="${x}">${x}</option>`).join("");
    if([...sa.options].some(o=>o.value===cur)) sa.value = cur;
    else sa.value = "All";

    // sync hidden values after filter updates
    syncMultiSelectHiddenValue(msMonth, month);
    syncMultiSelectHiddenValue(msDiv, division);
  }

  function clearPointLabels(){
    pointEls.forEach(el=>el.remove());
    pointEls = [];
  }

  function computeMonthWiseCC(rows){
    const map = new Map();
    FY_MONTHS.forEach(m=>map.set(m, {links:0, concern:0}));
    rows.forEach(r=>{
      const m = (r["Month"] || "").trim();
      if(!map.has(m)) map.set(m, {links:0, concern:0});
      const links = Number(r["Links Triggered"] || 0);
      const ccnt = Number(r["Concern Count"] || 0);
      const obj = map.get(m);
      obj.links += links;
      obj.concern += ccnt;
    });

    const labels = [];
    const values = [];
    FY_MONTHS.forEach(m=>{
      const obj = map.get(m);
      if(!obj) return;
      if(obj.links <= 0) return;
      const v = (obj.concern / obj.links) * 1000.0;
      labels.push(m);
      values.push(Number(v.toFixed(2)));
    });
    return {labels, values};
  }

  function drawPointLabels(labels, values){
    clearPointLabels();
    if(!ccChart) return;

    const meta = ccChart.getDatasetMeta(0);
    if(!meta || !meta.data) return;

    for(let i=0;i<meta.data.length;i++){
      const pt = meta.data[i];
      const v = values[i];
      if(v === null || v === undefined) continue;

      const el = document.createElement("div");
      el.className = "ptLabel";
      el.textContent = Number(v).toFixed(2);
      el.style.left = pt.x + "px";
      el.style.top = pt.y + "px";
      chartWrap.appendChild(el);
      pointEls.push(el);
    }
  }

  function renderChart(rows){
    const {labels, values} = computeMonthWiseCC(rows);

    if(!labels.length || !values.length){
      chartCard.classList.remove("show");
      clearPointLabels();
      if(ccChart){ ccChart.destroy(); ccChart = null; }
      return;
    }

    chartCard.classList.add("show");

    const avg = values.reduce((a,b)=>a+b,0) / values.length;
    const avgLine = labels.map(()=>avg);

    const data = {
      labels,
      datasets: [
        {
          label: "CC/1000",
          data: values,
          borderColor: "#f97316",
          backgroundColor: "rgba(249,115,22,0.20)",
          tension: 0.4,
          pointRadius: 5,
          pointHoverRadius: 7,
          fill: true
        },
        {
          label: "Average",
          data: avgLine,
          borderColor: "rgba(15,23,42,0.45)",
          borderDash: [6,6],
          pointRadius: 0,
          tension: 0,
          fill: false
        }
      ]
    };

    const options = {
      responsive: true,
      maintainAspectRatio: false,
      plugins: {
        legend: { display: true },
        tooltip: { enabled: true }
      },
      scales: {
        x: { grid: { display: false } },
        y: { beginAtZero: true }
      },
      animation: { onComplete: () => drawPointLabels(labels, values) }
    };

    if(ccChart){
      ccChart.data = data;
      ccChart.options = options;
      ccChart.update();
      setTimeout(()=>drawPointLabels(labels, values), 60);
    }else{
      ccChart = new Chart(ccChartCanvas.getContext("2d"), { type: "line", data, options });
    }
  }

  function colorClassPercent(n){ if(n===null) return ""; return (n >= 30) ? "pill-green" : "pill-red"; }
  function colorClassCC(n){ if(n===null) return ""; return (n > 20) ? "pill-red" : "pill-green"; }
  function colorClassOSAT(n){ if(n===null) return ""; return (n >= 70) ? "pill-green" : "pill-red"; }

  function renderTable(rows){
    tbody.innerHTML = "";
    const colspan = (TABLE_MODE === "bodyshop") ? 8 : 10;

    if(!rows || rows.length===0){
      tbody.innerHTML = `<tr><td colspan="${colspan}" style="text-align:center;color:#64748b;padding:16px;">No records found</td></tr>`;
      count.querySelector("span").textContent = "Showing 0 records";
      return;
    }

    rows.forEach(r=>{
      const p = num(r["% of Response"]);
      const cc = num(r["CC/1000"]);
      const os = num(r["OSAT"]);
      const divDisp = divShort(r["Division"]);

      if(TABLE_MODE === "bodyshop"){
        tbody.insertAdjacentHTML("beforeend", `
          <tr>
            <td>${safe(r["Month"])}</td>
            <td><b>${safe(r["SA Name"])}</b></td>
            <td>${safe(r["Links Triggered"])}</td>
            <td>${safe(r["Response"])}</td>
            <td class="${colorClassPercent(p)}">${pctText(p)}</td>
            <td>${safe(r["Concern Count"])}</td>
            <td class="${colorClassCC(cc)}">${safe(r["CC/1000"])}</td>
            <td>${safe(divDisp)}</td>
          </tr>
        `);
      }else{
        tbody.insertAdjacentHTML("beforeend", `
          <tr>
            <td>${safe(r["Month"])}</td>
            <td><b>${safe(r["SA Name"])}</b></td>
            <td>${safe(r["Links Triggered"])}</td>
            <td>${safe(r["Response"])}</td>
            <td class="${colorClassPercent(p)}">${pctText(p)}</td>
            <td>${safe(r["Concern Count"])}</td>
            <td class="${colorClassCC(cc)}">${safe(r["CC/1000"])}</td>
            <td class="${colorClassOSAT(os)}">${safe(r["OSAT"])}</td>
            <td>${safe(r["NPS"])}</td>
            <td>${safe(divDisp)}</td>
          </tr>
        `);
      }
    });

    count.querySelector("span").textContent = "Showing " + rows.length + " records";
  }

  async function refresh(){
    showLoading(true);
    err.classList.remove("show");
    try{
      const res = await fetch("/api/data?" + params());
      const j = await res.json();
      if(!j.ok) throw new Error(j.error || "Failed to load data");

      allRows = j.rows || [];
      const sum = j.summary || {};

      k1.textContent = sum.total_links_triggered ?? "-";
      k2.textContent = sum.total_responses ?? "-";
      k3.textContent = (sum.avg_percent_response==null) ? "-" : Number(sum.avg_percent_response).toFixed(2) + "%";
      k4.textContent = sum.total_concern_count ?? "-";
      k5.textContent = (sum.avg_cc_per_1000==null) ? "-" : Number(sum.avg_cc_per_1000).toFixed(2);
      k6.textContent = SHOW_OSAT ? ((sum.avg_osat==null) ? "-" : Number(sum.avg_osat).toFixed(2)) : "-";

      // paging
      const ps = Number(pageSize.value || 50);
      const start = (page-1)*ps;
      const pageRows = allRows.slice(start, start+ps);

      renderTable(pageRows);
      renderChart(allRows);
    }catch(e){
      err.textContent = String(e.message || e);
      err.classList.add("show");
    }finally{
      showLoading(false);
    }
  }

  document.getElementById("reset").addEventListener("click", async ()=>{
    // reset: select all month & div
    msMonth.selected.clear(); msMonth.options.forEach(o=>msMonth.selected.add(o));
    msDiv.selected.clear(); msDiv.options.forEach(o=>msDiv.selected.add(o));
    syncMultiSelectUI(msMonth); syncMultiSelectUI(msDiv);
    syncMultiSelectHiddenValue(msMonth, month);
    syncMultiSelectHiddenValue(msDiv, division);
    sa.value = "All";
    page = 1;
    await loadFilters();
    await refresh();
  });

  document.getElementById("prevBtn").addEventListener("click", async ()=>{
    if(page<=1) return;
    page--;
    await refresh();
  });
  document.getElementById("nextBtn").addEventListener("click", async ()=>{
    const ps = Number(pageSize.value || 50);
    const maxPage = Math.max(1, Math.ceil((allRows.length||0)/ps));
    if(page>=maxPage) return;
    page++;
    await refresh();
  });
  pageSize.addEventListener("change", async ()=>{
    page = 1;
    await refresh();
  });

  document.getElementById("exportBtn").addEventListener("click", ()=>{
    window.open("/api/export?" + params(), "_blank");
  });

  // init
  (async function init(){
    await loadFilters();
    await refresh();
  })();
</script>
</body>
</html>
"""

# =============================================================================
# 7) PAGES
# =============================================================================
def render_page(key: str):
    meta = DATASET_META.get(key, DATASET_META["personal"])
    return render_template_string(
        INDEX_HTML,
        page_title=meta["page_title"],
        active_key=key,
        show_osat=bool(meta.get("show_osat", True)),
        table_mode=meta.get("table_mode", "normal"),
    )


@app.route("/", methods=["GET"])
def home():
    return render_page("personal")


@app.route("/personal", methods=["GET"])
def personal_page():
    return render_page("personal")


@app.route("/meal", methods=["GET"])
def meal_page():
    return render_page("meal")


@app.route("/bodyshop", methods=["GET"])
def bodyshop_page():
    return render_page("bodyshop")


@app.route("/commercial", methods=["GET"])
def commercial_page():
    return render_page("commercial")


# =============================================================================
# 8) APIs
# =============================================================================
@app.route("/api/filters", methods=["GET"])
def api_filters():
    ds_key = request.args.get("dataset", "personal").strip().lower()
    ds = get_dataset(ds_key)

    if ds.load_error:
        return jsonify({"ok": False, "error": f"{ds.name}: {ds.load_error}"}), 500

    month = request.args.get("month", "All")
    division = request.args.get("division", "All")

    filters = ds.get_filters(month, division)
    return jsonify({"ok": True, "filters": filters})


@app.route("/api/data", methods=["GET"])
def api_data():
    ds_key = request.args.get("dataset", "personal").strip().lower()
    ds = get_dataset(ds_key)
    meta = DATASET_META.get(ds_key, DATASET_META["personal"])

    if ds.load_error:
        return jsonify({"ok": False, "error": f"{ds.name}: {ds.load_error}"}), 500

    month = request.args.get("month", "All")
    division = request.args.get("division", "All")
    sa_name = request.args.get("sa_name", "All")

    rows = ds.apply_filters(month, division, sa_name)
    summary = ds.compute_kpis(rows, include_osat=bool(meta.get("show_osat", True)))
    return jsonify({"ok": True, "rows": rows, "summary": summary})


@app.route("/api/summary", methods=["GET"])
def api_summary():
    ds_key = request.args.get("dataset", "personal").strip().lower()
    ds = get_dataset(ds_key)
    meta = DATASET_META.get(ds_key, DATASET_META["personal"])

    if ds.load_error:
        return jsonify({"ok": False, "error": f"{ds.name}: {ds.load_error}"}), 500

    month = request.args.get("month", "All")
    division = request.args.get("division", "All")
    sa_name = request.args.get("sa_name", "All")

    rows = ds.apply_filters(month, division, sa_name)
    summary = ds.compute_kpis(rows, include_osat=bool(meta.get("show_osat", True)))
    return jsonify({"ok": True, "summary": summary})


@app.route("/api/export", methods=["GET"])
def api_export():
    ds_key = request.args.get("dataset", "personal")
    ds = get_dataset(ds_key)
    meta = DATASET_META.get((ds_key or "").strip().lower(), DATASET_META["personal"])

    if ds.load_error:
        return jsonify({"ok": False, "error": f"{ds.name}: {ds.load_error}"}), 500

    month = request.args.get("month", "All")
    division = request.args.get("division", "All")
    sa_name = request.args.get("sa_name", "All")

    rows = ds.apply_filters(month, division, sa_name)

    cols = meta.get("export_cols") or [
        "Month",
        "SA Name",
        "Links Triggered",
        "Response",
        "NPS",
        "% of Response",
        "Concern Count",
        "CC/1000",
        "OSAT",
        "Division",
        "Mile id",
    ]

    bio = export_xlsx(cols, rows)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"Service_Intello_{ds.name}_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =============================================================================
# 9) LOCAL AUTO OPEN (OPTIONAL)
# =============================================================================
def open_browser():
    webbrowser.open_new(f"http://{HOST}:{PORT}/")


# =============================================================================
# 10) MAIN (LOCAL RUN ONLY)
# =============================================================================
if __name__ == "__main__":
    if AUTO_OPEN_BROWSER:
        threading.Timer(1.0, open_browser).start()

    app.run(host=HOST, port=PORT, debug=True, use_reloader=False)
